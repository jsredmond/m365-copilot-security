# Generate Copilot Readiness Assessment Excel Template
# Requires: openpyxl library

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

def create_assessment_workbook():
    """
    Creates comprehensive Excel assessment template for Copilot readiness
    """
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    
    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    summary_data = [
        ["Copilot Security Readiness Assessment", "", "", ""],
        ["", "", "", ""],
        ["Assessment Date:", datetime.now().strftime("%Y-%m-%d"), "", ""],
        ["Organization:", "Enter Organization Name", "", ""],
        ["Assessed By:", "Enter Name", "", ""],
        ["", "", "", ""],
        ["READINESS SCORES", "", "", ""],
        ["Category", "Current Score", "Target Score", "Gap"],
        ["Data Governance", 0, 85, "=C8-B8"],
        ["Access Controls", 0, 90, "=C9-B9"],
        ["Sensitivity Labels", 0, 80, "=C10-B10"],
        ["DLP Policies", 0, 95, "=C11-B11"],
        ["Audit & Monitoring", 0, 85, "=C12-B12"],
        ["Overall Readiness", "=AVERAGE(B8:B12)", 87, "=C13-B13"]
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Format headers
    for cell in ws_summary[7]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Sheet 2: Detailed Checklist
    ws_checklist = wb.create_sheet("Security Checklist")
    
    checklist_headers = [
        "Category", "Control", "Status", "Priority", 
        "Owner", "Due Date", "Notes", "Evidence"
    ]
    
    checklist_items = [
        ["Data Discovery", "Complete SharePoint audit", "Not Started", "Critical", "", "", "", ""],
        ["Data Discovery", "Identify stale sites", "Not Started", "High", "", "", "", ""],
        ["Data Discovery", "Review external sharing", "Not Started", "Critical", "", "", "", ""],
        ["Access Control", "Enable Restricted Search", "Not Started", "Critical", "", "", "", ""],
        ["Access Control", "Configure RAC policies", "Not Started", "High", "", "", "", ""],
        ["Sensitivity Labels", "Create label taxonomy", "Not Started", "Critical", "", "", "", ""],
        ["Sensitivity Labels", "Deploy auto-labeling", "Not Started", "Medium", "", "", "", ""],
        ["DLP", "Create Copilot DLP policy", "Not Started", "Critical", "", "", "", ""],
        ["DLP", "Test false positives", "Not Started", "High", "", "", "", ""],
        ["Monitoring", "Deploy Sentinel workbooks", "Not Started", "High", "", "", "", ""],
        ["Monitoring", "Configure alerts", "Not Started", "High", "", "", "", ""],
    ]
    
    ws_checklist.append(checklist_headers)
    for item in checklist_items:
        ws_checklist.append(item)
    
    # Sheet 3: Risk Register
    ws_risks = wb.create_sheet("Risk Register")
    
    risk_headers = [
        "Risk ID", "Category", "Description", "Likelihood", 
        "Impact", "Risk Score", "Mitigation", "Status"
    ]
    
    risks = [
        ["R001", "Data Exposure", "Overshared sites exposed via Copilot", "High", "High", "=D2*E2", "Implement Restricted Search", "Open"],
        ["R002", "Compliance", "PII/PHI in Copilot responses", "Medium", "Critical", "=D3*E3", "Deploy DLP policies", "Open"],
        ["R003", "Access Control", "Unauthorized access to sensitive data", "Medium", "High", "=D4*E4", "Enforce sensitivity labels", "Open"],
    ]
    
    ws_risks.append(risk_headers)
    for risk in risks:
        ws_risks.append(risk)
    
    # Sheet 4: Metrics Tracking
    ws_metrics = wb.create_sheet("Metrics")
    
    metrics_headers = [
        "Week", "External Sharing Sites", "Labeled Content %", 
        "DLP Violations", "Guest Users", "Copilot Users", "Incidents"
    ]
    
    # Add headers and sample week entries
    ws_metrics.append(metrics_headers)
    for week in range(1, 17):  # 16-week timeline
        ws_metrics.append([f"Week {week}", "", "", "", "", "", ""])
    
    # Sheet 5: License Planning
    ws_licenses = wb.create_sheet("License Planning")
    
    license_headers = [
        "Department", "Users", "Current License", "Required License", 
        "Copilot License", "Cost/Month", "Notes"
    ]
    
    departments = [
        ["Finance", 50, "E3", "E5", "Yes", "=B2*65", "Pilot group"],
        ["Sales", 100, "E3", "E5", "Yes", "=B3*65", "Phase 2"],
        ["IT", 25, "E5", "E5", "Yes", "=B4*30", "Pilot group"],
        ["HR", 30, "E3", "E5", "Planned", "=B5*65", "Phase 3"],
    ]
    
    ws_licenses.append(license_headers)
    for dept in departments:
        ws_licenses.append(dept)
    
    # Save workbook
    filename = f"Copilot_Readiness_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    return filename

# Generate the template
template_file = create_assessment_workbook()
print(f"Assessment template created: {template_file}")